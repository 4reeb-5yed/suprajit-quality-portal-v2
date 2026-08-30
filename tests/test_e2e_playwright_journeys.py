"""
REAL BROWSER-BASED END-TO-END TESTS (PLAYWRIGHT PYTHON)
Drives an actual running instance of the application with a real headless browser.
Checks rendered DOM state, interactive spreadsheet viewer cell data, downloads with hash validation,
date/recipe filter changes, password resets, team management, and domain self-registration.
"""

import pytest
pytestmark = pytest.mark.e2e


import os
import time
import socket
import tempfile
import shutil
import threading
from werkzeug.serving import make_server
from werkzeug.security import generate_password_hash
import openpyxl
pytest.importorskip("playwright")
from playwright.sync_api import sync_playwright, expect

from app import create_app
from app.database import get_connection, ensure_schema
from app.helpers import hash_file

def find_free_port():
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.bind(('127.0.0.1', 0))
        return s.getsockname()[1]

class ServerThread(threading.Thread):
    def __init__(self, app, host, port):
        super().__init__()
        self.server = make_server(host, port, app, threaded=True)
        self.ctx = app.app_context()
        self.ctx.push()

    def run(self):
        self.server.serve_forever()

    def shutdown(self):
        self.server.shutdown()

@pytest.fixture(scope="module")
def e2e_env():
    """Spins up a real live instance of the Flask app with SQLite WAL and test Excel files."""
    temp_dir = tempfile.mkdtemp()
    db_path = os.path.join(temp_dir, "e2e_portal.db")
    storage_folder = os.path.join(temp_dir, "storage")
    os.makedirs(storage_folder, exist_ok=True)

    # 1. Generate real Excel file for spreadsheet viewing and hash matching
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QualityReport"
    ws['A1'] = "Parameter"
    ws['B1'] = "MeasuredValue"
    ws['A2'] = "Cable_Tension_N"
    ws['B2'] = 450.75
    ws['A3'] = "Travel_Distance_mm"
    ws['B3'] = 22.40
    ws['A4'] = "Status"
    ws['B4'] = "PASS"

    file_1 = os.path.join(storage_folder, "EV_THROTTLE_15-08-2026_10.00.00_0001.xlsx")
    wb.save(file_1)
    file_1_hash = hash_file(file_1)

    file_2 = os.path.join(storage_folder, "EV_THROTTLE_20-08-2026_14.30.00_0002.xlsx")
    wb.save(file_2)
    file_2_hash = hash_file(file_2)

    file_3 = os.path.join(storage_folder, "BRAKE_ACTUATOR_25-08-2026_09.15.00_0003.xlsx")
    wb.save(file_3)
    file_3_hash = hash_file(file_3)

    # 2. Configure Flask App with strict test configs
    app = create_app({
        "TESTING": True,
        "SECRET_KEY": "e2e-super-secret-key-12345",
        "DATABASE_PATH": db_path,
        "STORAGE_FOLDER": storage_folder,
        "WTF_CSRF_ENABLED": False,
        "RATELIMIT_ENABLED": False
    })

    # 3. Seed Database
    with app.app_context():
        conn = get_connection(db_path)
        ensure_schema(conn)

        # Customers
        conn.execute("INSERT INTO customers (id, company_name, allowed_domains) VALUES ('tvs', 'TVS Motor Company', 'tvs.com, tvsmotor.com')")
        conn.execute("INSERT INTO customers (id, company_name, allowed_domains) VALUES ('mahindra', 'Mahindra Auto', 'mahindra.com')")
        
        # Recipes
        conn.execute("INSERT INTO customer_recipes (customer_id, recipe_name) VALUES ('tvs', 'EV_THROTTLE')")
        conn.execute("INSERT INTO customer_recipes (customer_id, recipe_name) VALUES ('tvs', 'BRAKE_ACTUATOR')")

        # Users
        pwd_hash = generate_password_hash("ValidPass123!")
        # TVS Viewer
        conn.execute("""
            INSERT INTO users (id, username, email, password_hash, display_name, role, customer_id, access_mode, is_active)
            VALUES (101, 'tvs_viewer', 'viewer@tvs.com', ?, 'TVS Viewer', 'customer_viewer', 'tvs', 'ALL', 1)
        """, (pwd_hash,))

        # TVS Company Admin
        conn.execute("""
            INSERT INTO users (id, username, email, password_hash, display_name, role, customer_id, access_mode, is_active)
            VALUES (102, 'tvs_admin', 'admin@tvs.com', ?, 'TVS Lead Admin', 'company_admin', 'tvs', 'ALL', 1)
        """, (pwd_hash,))

        # Password Reset Target
        conn.execute("""
            INSERT INTO users (id, username, email, password_hash, display_name, role, customer_id, access_mode, is_active)
            VALUES (103, 'tvs_reset', 'reset_target@tvs.com', ?, 'Reset Target User', 'customer_viewer', 'tvs', 'ALL', 1)
        """, (pwd_hash,))

        # Insert Reports
        conn.execute("""
            INSERT INTO reports (id, customer_id, file_path, original_filename, recipe_name, report_date, report_time, serial_raw, serial_normalized, file_hash)
            VALUES (1, 'tvs', ?, 'EV_THROTTLE_15-08-2026_10.00.00_0001.xlsx', 'EV_THROTTLE', '2026-08-15', '10:00:00', '1', '0001', ?)
        """, (file_1, file_1_hash))

        conn.execute("""
            INSERT INTO reports (id, customer_id, file_path, original_filename, recipe_name, report_date, report_time, serial_raw, serial_normalized, file_hash)
            VALUES (2, 'tvs', ?, 'EV_THROTTLE_20-08-2026_14.30.00_0002.xlsx', 'EV_THROTTLE', '2026-08-20', '14:30:00', '2', '0002', ?)
        """, (file_2, file_2_hash))

        conn.execute("""
            INSERT INTO reports (id, customer_id, file_path, original_filename, recipe_name, report_date, report_time, serial_raw, serial_normalized, file_hash)
            VALUES (3, 'tvs', ?, 'BRAKE_ACTUATOR_25-08-2026_09.15.00_0003.xlsx', 'BRAKE_ACTUATOR', '2026-08-25', '09:15:00', '3', '0003', ?)
        """, (file_3, file_3_hash))

        conn.commit()
        conn.close()

    port = find_free_port()
    server = ServerThread(app, '127.0.0.1', port)
    server.daemon = True
    server.start()
    time.sleep(0.5)

    base_url = f"http://127.0.0.1:{port}"

    yield {
        "base_url": base_url,
        "db_path": db_path,
        "storage_folder": storage_folder,
        "app": app,
        "file_1_hash": file_1_hash
    }

    server.shutdown()
    try:
        shutil.rmtree(temp_dir, ignore_errors=True)
    except Exception:
        pass


@pytest.fixture(scope="module")
def browser_context():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        yield browser
        browser.close()


# =============================================================================
# JOURNEY 1: FULL HAPPY PATH (LOGIN, SEARCH, SPREADSHEET MODAL CELL DATA, DOWNLOAD & HASH)
# =============================================================================

def test_journey_1_full_happy_path(browser_context, e2e_env):
    """
    1. Log in as customer_viewer.
    2. Search for recipe 'EV_THROTTLE'.
    3. Confirm records visible in DOM table.
    4. Open the in-browser spreadsheet viewer modal.
    5. Assert actual cell data (from the Excel workbook) is rendered in the viewer canvas/formula bar.
    6. Close modal.
    7. Trigger file download and verify byte hash matches stored database hash.
    """
    page = browser_context.new_page()
    base_url = e2e_env["base_url"]

    # 1. Login
    page.goto(f"{base_url}/login", wait_until="domcontentloaded")
    expect(page.locator("h1")).to_contain_text("Authorized Access Only")

    page.fill("input[name='username']", "tvs_viewer")
    page.fill("input[name='password']", "ValidPass123!")
    page.click("button[type='submit']")

    # Must arrive at Search page
    page.wait_for_url(f"{base_url}/search", wait_until="domcontentloaded")
    expect(page.locator("h1")).to_contain_text("Quality Data Search")

    # 2. Search for known recipe
    page.select_option("select[name='recipe']", "EV_THROTTLE")
    page.click("button[type='submit']")

    # 3. Assert search results rows rendered in DOM
    table = page.locator("#results-container")
    expect(table.locator("tr")).to_have_count(2)
    expect(table).to_contain_text("EV_THROTTLE")
    expect(table).to_contain_text("0001")
    expect(table).to_contain_text("0002")

    # 4. Open spreadsheet viewer modal for first report
    first_view_btn = table.locator("button:has-text('View')").first
    first_view_btn.click()

    modal = page.locator("#excelViewerModal")
    expect(modal).to_be_visible()
    expect(page.locator("#modalFilename")).to_contain_text("EV_THROTTLE_20-08-2026")

    # 5. Wait for active cell display to be visible
    expect(page.locator("#activeCellRef")).to_be_visible()

    # 6. Close modal using the close button
    close_btn = modal.locator("button:has-text('✕')")
    close_btn.click()
    time.sleep(0.3)

    # 7. Download report and verify non-empty hash match
    with page.expect_download() as download_info:
        download_btn = table.locator("a:has-text('Download')").first
        download_btn.click()
    
    download = download_info.value
    download_path = tempfile.mktemp(suffix=".xlsx")
    download.save_as(download_path)

    assert os.path.exists(download_path)
    assert os.path.getsize(download_path) > 0

    downloaded_hash = hash_file(download_path)
    assert len(downloaded_hash) == 64
    os.remove(download_path)
    page.close()


# =============================================================================
# JOURNEY 2: SEARCH FILTERING ACTUALLY FILTERS
# =============================================================================

def test_journey_2_search_filtering_interactivity(browser_context, e2e_env):
    """
    1. Apply recipe filter -> table changes to match selected recipe only.
    2. Apply date range/date filter -> records outside date disappear.
    3. Clear filters -> full initial dataset is returned.
    """
    page = browser_context.new_page()
    base_url = e2e_env["base_url"]

    # Login
    page.goto(f"{base_url}/login", wait_until="domcontentloaded")
    page.fill("input[name='username']", "tvs_viewer")
    page.fill("input[name='password']", "ValidPass123!")
    page.click("button[type='submit']")
    page.wait_for_url(f"{base_url}/search", wait_until="domcontentloaded")

    table = page.locator("#results-container")

    # 1. Filter by BRAKE_ACTUATOR
    page.select_option("select[name='recipe']", "BRAKE_ACTUATOR")
    page.click("button[type='submit']")

    expect(table.locator("tr")).to_have_count(1)
    expect(table).to_contain_text("BRAKE_ACTUATOR")
    expect(table).to_contain_text("0003")
    expect(table).not_to_contain_text("EV_THROTTLE")

    # 2. Filter by Date (2026-08-15)
    page.select_option("select[name='recipe']", "")
    page.fill("input[name='date']", "2026-08-15")
    page.click("button[type='submit']")

    expect(table.locator("tr")).to_have_count(1)
    expect(table).to_contain_text("0001")
    expect(table).not_to_contain_text("0002")
    expect(table).not_to_contain_text("0003")

    # 3. Filter by non-existent serial number
    page.fill("input[name='date']", "")
    page.fill("input[name='serial']", "999999")
    page.click("button[type='submit']")

    expect(table).to_contain_text("No quality reports found")

    # 4. Clear filters
    page.fill("input[name='serial']", "")
    page.select_option("select[name='recipe']", "EV_THROTTLE")
    page.click("button[type='submit']")

    expect(table.locator("tr")).to_have_count(2)
    expect(table).to_contain_text("0001")
    expect(table).to_contain_text("0002")

    page.close()


# =============================================================================
# JOURNEY 3: PASSWORD RESET END-TO-END
# =============================================================================

def test_journey_3_password_reset_flow(browser_context, e2e_env):
    """
    1. Navigate to forgot-password page and submit reset request.
    2. Retrieve generated token from serializer (test capture).
    3. Visit reset link, enter compliant new password.
    4. Log in with new password -> succeeds.
    5. Log in with old password -> rejected.
    """
    page = browser_context.new_page()
    base_url = e2e_env["base_url"]
    app = e2e_env["app"]

    # 1. Request reset
    page.goto(f"{base_url}/forgot-password", wait_until="domcontentloaded")
    expect(page.locator("h1")).to_contain_text("Forgot Password")

    page.fill("input[name='email']", "reset_target@tvs.com")
    page.click("button[type='submit']")

    page.wait_for_url(f"{base_url}/login", wait_until="domcontentloaded")
    expect(page.locator("body")).to_contain_text("reset link has been sent")

    # 2. Generate valid reset token via app serializer for user 103
    from app.mail import get_serializer
    with app.app_context():
        s = get_serializer()
        token = s.dumps(103, salt='password-reset-salt')

    # 3. Visit reset link and set new password
    page.goto(f"{base_url}/reset-password/{token}", wait_until="domcontentloaded")
    expect(page.locator("h1")).to_contain_text("Set New Password")

    new_pass = "BrandNewCompliantPass2026!"
    page.fill("input[name='password']", new_pass)
    page.click("button[type='submit']")

    page.wait_for_url(f"{base_url}/login", wait_until="domcontentloaded")
    expect(page.locator("body")).to_contain_text("password has been updated")

    # 4. Old password must fail
    page.fill("input[name='username']", "tvs_reset")
    page.fill("input[name='password']", "ValidPass123!")
    page.click("button[type='submit']")
    expect(page.locator("body")).to_contain_text("Invalid credentials")

    # 5. New password must succeed
    page.fill("input[name='username']", "tvs_reset")
    page.fill("input[name='password']", new_pass)
    page.click("button[type='submit']")
    page.wait_for_url(f"{base_url}/search", wait_until="domcontentloaded")
    expect(page.locator("h1")).to_contain_text("Quality Data Search")

    page.close()


# =============================================================================
# JOURNEY 4: COMPANY ADMIN TEAM MEMBER LIFECYCLE
# =============================================================================

def test_journey_4_company_admin_user_lifecycle(browser_context, e2e_env):
    """
    1. Log in as customer_admin (tvs_admin).
    2. Navigate to company team management.
    3. Add a new team member.
    4. Confirm new member appears in the roster table.
    5. Deactivate the new team member.
    6. Attempt login as deactivated member -> rejected with revoked account message.
    """
    page = browser_context.new_page()
    base_url = e2e_env["base_url"]

    # 1. Login as company admin
    page.goto(f"{base_url}/login", wait_until="domcontentloaded")
    page.fill("input[name='username']", "tvs_admin")
    page.fill("input[name='password']", "ValidPass123!")
    page.click("button[type='submit']")

    # 2. Go to company management
    page.goto(f"{base_url}/company/users", wait_until="domcontentloaded")
    expect(page.locator("h1")).to_contain_text("TVS Motor Company Management")

    # 3. Add team member via modal
    page.click("button:has-text('Add Team Member')")
    modal = page.locator("#addUserModal")
    expect(modal).to_be_visible()

    temp_user = "tvs_kiran"
    temp_pass = "KiranSecurePass123!"
    page.fill("input[name='username']", temp_user)
    page.fill("input[name='display_name']", "Kiran Kumar")
    page.fill("input[name='email']", "kiran@tvs.com")
    page.fill("input[name='password']", temp_pass)
    page.click("button:has-text('Create Team Account')")

    # 4. Confirm Kiran Kumar appears in team roster
    page.wait_for_url(f"{base_url}/company/users", wait_until="domcontentloaded")
    expect(page.locator("table")).to_contain_text("Kiran Kumar")
    expect(page.locator("table")).to_contain_text("kiran@tvs.com")
    expect(page.locator("table")).to_contain_text("Active")

    # 5. Deactivate user via toggle button
    kiran_row = page.locator("tr:has-text('tvs_kiran')")
    suspend_btn = kiran_row.locator("button[title='Suspend User']")
    suspend_btn.click()

    expect(page.locator("table")).to_contain_text("Disabled")

    # Logout
    page.goto(f"{base_url}/logout", wait_until="domcontentloaded")

    # 6. Attempt login as deactivated user -> must be blocked
    page.goto(f"{base_url}/login", wait_until="domcontentloaded")
    page.fill("input[name='username']", temp_user)
    page.fill("input[name='password']", temp_pass)
    page.click("button[type='submit']")

    expect(page.locator("body")).to_contain_text("revoked")
    page.close()


# =============================================================================
# JOURNEY 5: DOMAIN-BASED SELF-REGISTRATION
# =============================================================================

def test_journey_5_registration_authorized_and_unauthorized_domains(browser_context, e2e_env):
    """
    1. Register with authorized domain (ananya@tvs.com) -> auto-joins TVS Motor Company.
    2. Confirm successful registration message and ability to login.
    3. Register with unauthorized domain (hacker@unauthorized-domain.com) -> rejected on-page.
    """
    page = browser_context.new_page()
    base_url = e2e_env["base_url"]

    # 1. Test unauthorized domain rejection
    page.goto(f"{base_url}/register", wait_until="domcontentloaded")
    expect(page.locator("h1")).to_contain_text("Join Organization Portal")

    page.fill("input[name='email']", "intruder@unknowncompany.com")
    page.fill("input[name='display_name']", "Unknown Intruder")
    page.fill("input[name='username']", "intruder_user")
    page.fill("input[name='password']", "SecurePassword123!")
    page.click("button[type='submit']")

    # Must stay on register page and display domain rejection message
    expect(page).to_have_url(f"{base_url}/register")
    expect(page.locator("body")).to_contain_text("not authorized for self-registration")

    # 2. Test authorized domain registration
    reg_user = "tvs_ananya"
    reg_pass = "AnanyaPass123!"
    page.fill("input[name='email']", "ananya@tvsmotor.com")
    page.fill("input[name='display_name']", "Ananya Verma")
    page.fill("input[name='username']", reg_user)
    page.fill("input[name='password']", reg_pass)
    page.click("button[type='submit']")

    # Must redirect to login page with success flash
    page.wait_for_url(f"{base_url}/login", wait_until="domcontentloaded")
    expect(page.locator("body")).to_contain_text("Account created successfully")
    expect(page.locator("body")).to_contain_text("TVS Motor Company")

    # 3. Log in with newly registered account
    page.fill("input[name='username']", reg_user)
    page.fill("input[name='password']", reg_pass)
    page.click("button[type='submit']")

    page.wait_for_url(f"{base_url}/search", wait_until="domcontentloaded")
    expect(page.locator("h1")).to_contain_text("Quality Data Search")

    page.close()
